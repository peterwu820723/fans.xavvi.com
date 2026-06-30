from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl


def cell_ref(image) -> str | None:
    anchor = getattr(image, "anchor", None)
    marker = getattr(anchor, "_from", None)
    if marker is None:
        return None
    return f"{openpyxl.utils.get_column_letter(marker.col + 1)}{marker.row + 1}"


def main() -> None:
    if len(sys.argv) != 2:
        raise SystemExit("Usage: inspect_workbook.py <workbook.xlsx>")

    workbook_path = Path(sys.argv[1])
    workbook = openpyxl.load_workbook(workbook_path, data_only=False)
    summary = {
        "path": str(workbook_path),
        "sheets": [],
    }

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        non_empty_rows = [
            [value for value in row]
            for row in rows
            if any(value is not None and str(value).strip() for value in row)
        ]
        images = []
        for index, image in enumerate(getattr(sheet, "_images", []), start=1):
            images.append(
                {
                    "index": index,
                    "anchor": cell_ref(image),
                    "format": getattr(image, "format", None),
                    "width": getattr(image, "width", None),
                    "height": getattr(image, "height", None),
                }
            )

        summary["sheets"].append(
            {
                "title": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "image_count": len(images),
                "images": images[:30],
                "sample_rows": non_empty_rows[:8],
            }
        )

    print(json.dumps(summary, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
